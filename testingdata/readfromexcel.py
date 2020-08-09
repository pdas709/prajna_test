import openpyxl
from random import randint

class ReadWriteExcel:

  def readandADDurl(self):
    path = "D:\\pytest_project\\testingdata\\Pytest.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 2,column = 1)
    return cell_obj.value

  def username(self):
    path = "D:\\pytest_project\\testingdata\\Pytest.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 2,column = 2)
    return cell_obj.value
  def password(self):
    path = "D:\\pytest_project\\testingdata\\Pytest.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 2,column = 3)
    return cell_obj.value

